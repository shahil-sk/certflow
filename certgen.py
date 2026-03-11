import os
import sys
import re
import json
import threading
import platform
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
except ImportError:
    messagebox.showerror("Missing Dependency", "Pillow is not installed.\nRun: pip install pillow")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    messagebox.showerror("Missing Dependency", "openpyxl is not installed.\nRun: pip install openpyxl")
    sys.exit(1)

try:
    from fpdf import FPDF
except ImportError:
    messagebox.showerror("Missing Dependency", "fpdf2 is not installed.\nRun: pip install fpdf2")
    sys.exit(1)


def resource_path(relative_path: str) -> str:
    """Get absolute path for bundled or dev resources."""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)


class CertificateApp:

    APP_TITLE = "CertWizard - Professional Certificate Generator"
    BG_DARK   = "#2c3e50"
    BG_LIGHT  = "#f0f2f5"
    ACCENT    = "#3498db"
    GREEN     = "#27ae60"
    WHITE     = "#ffffff"
    TEXT      = "#2c3e50"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(self.APP_TITLE)
        self.root.minsize(900, 600)

        # ── state ──────────────────────────────────────────────
        self.original_image: Image.Image | None = None
        self.display_image  = None
        self.scale_x = self.scale_y = 1.0
        self.template_path: str | None = None
        self.excel_path:    str | None = None
        self.placeholders:    dict = {}
        self._placeholder_images: dict = {}
        self._drag_data:     dict = {}
        self.excel_data:     list = []
        self.fields:         list = []
        self.field_vars:     dict = {}
        self.font_settings:  dict = {}
        self.color_space = tk.StringVar(value="RGB")
        self._gen_lock = threading.Lock()          # prevent double-generation

        self.available_fonts = self._load_fonts()
        self._setup_ui()
        self._set_icon()

    # ══════════════════════════════════════════════════════════════
    # UI SETUP
    # ══════════════════════════════════════════════════════════════

    def _setup_ui(self):
        self.root.configure(padx=15, pady=15, bg=self.BG_LIGHT)

        # ── navbar ─────────────────────────────────────────────
        nav = tk.Frame(self.root, bg=self.BG_DARK, height=50)
        nav.pack(fill="x", pady=(0, 10))

        tk.Label(nav, text="CertWizard", font=("Arial", 16, "bold"),
                 fg="white", bg=self.BG_DARK, padx=15).pack(side="left")

        proj_btn = tk.Menubutton(nav, text="Project", bg=self.BG_DARK, fg="white",
                                 relief="flat", font=("Arial", 10), padx=10)
        proj_btn.menu = tk.Menu(proj_btn, tearoff=0, bg=self.WHITE, fg=self.TEXT,
                                activebackground=self.ACCENT, activeforeground="white")
        proj_btn["menu"] = proj_btn.menu
        proj_btn.menu.add_command(label="Save Project", command=self.save_project)
        proj_btn.menu.add_command(label="Load Project", command=self.load_project)
        proj_btn.pack(side="left", padx=5)

        for label, cmd in (("Load Template", self.load_template),
                           ("Load Excel",    self.load_excel)):
            tk.Button(nav, text=label, command=cmd,
                      fg="white", bg=self.ACCENT, relief="flat",
                      padx=12, pady=5, font=("Arial", 9),
                      activebackground="#2980b9").pack(side="left", padx=5)

        # ── status bar ─────────────────────────────────────────
        self.status_bar = tk.Label(self.root, text="Ready",
                                   bd=0, relief=tk.FLAT, anchor=tk.W,
                                   bg=self.BG_DARK, fg="white", padx=10, pady=5)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # ── main area ──────────────────────────────────────────
        main = tk.Frame(self.root, bg=self.BG_LIGHT)
        main.pack(fill="both", expand=True)

        # ── left panel ─────────────────────────────────────────
        panel = tk.Frame(main, width=290, bg=self.WHITE, padx=15, pady=15)
        panel.pack(side="left", fill="y", padx=(0, 15))
        panel.pack_propagate(False)

        self.toggle_frame = tk.LabelFrame(
            panel, text="Certificate Fields",
            padx=10, pady=10, font=("Arial", 11, "bold"),
            bg=self.WHITE, fg=self.TEXT)
        self.toggle_frame.pack(fill="x", pady=(0, 10))

        btn_row = tk.Frame(panel, bg=self.WHITE)
        btn_row.pack(fill="x", pady=(10, 5))
        for label, cmd, col, active in (
                ("Preview",  self.preview_certificate,   self.GREEN, "#219a52"),
                ("Generate", self.generate_certificates, self.ACCENT, "#2980b9")):
            tk.Button(btn_row, text=label, command=cmd,
                      bg=col, fg="white", relief="flat",
                      padx=15, pady=8, font=("Arial", 10, "bold"),
                      activebackground=active
                      ).pack(side="left" if label == "Preview" else "right",
                             padx=2, fill="x", expand=True)

        style = ttk.Style()
        style.configure("Cert.Horizontal.TProgressbar",
                         troughcolor=self.BG_LIGHT, background=self.ACCENT, thickness=10)
        self.progress = ttk.Progressbar(panel, orient="horizontal", length=200,
                                        mode="determinate",
                                        style="Cert.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(10, 0))

        log_frame = tk.LabelFrame(panel, text="Generation Status",
                                  padx=10, pady=10, font=("Arial", 11, "bold"),
                                  bg=self.WHITE, fg=self.TEXT)
        log_frame.pack(fill="both", expand=True, pady=(10, 0))
        self.info_text = tk.Text(log_frame, height=8, width=30, wrap=tk.WORD,
                                 font=("Arial", 9), bg="#f8f9fa", fg=self.TEXT,
                                 relief=tk.FLAT, padx=5, pady=5, state="disabled")
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(log_frame, orient="vertical", command=self.info_text.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.info_text.configure(yscrollcommand=sb.set)

        # ── canvas ─────────────────────────────────────────────
        cv_frame = tk.Frame(main, relief="solid", borderwidth=1,
                            bg=self.WHITE, highlightbackground="#bdc3c7")
        cv_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        self.canvas = tk.Canvas(cv_frame, bg=self.WHITE, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=2, pady=2)

    # ══════════════════════════════════════════════════════════════
    # ICON
    # ══════════════════════════════════════════════════════════════

    def _set_icon(self):
        candidates = [
            resource_path("certgen.ico"),
            resource_path("certgen.png"),
            resource_path("icon.ico"),
            resource_path("icon.png"),
        ]
        for p in candidates:
            if not os.path.exists(p):
                continue
            try:
                if platform.system() == "Windows" and p.endswith(".ico"):
                    self.root.iconbitmap(p)
                else:
                    img   = Image.open(p)
                    photo = ImageTk.PhotoImage(img)
                    self.root.iconphoto(True, photo)
                    self._icon_ref = photo   # keep reference alive
                return
            except Exception as exc:
                print(f"Icon load failed ({p}): {exc}")

    # ══════════════════════════════════════════════════════════════
    # FONT HELPERS
    # ══════════════════════════════════════════════════════════════

    def _load_fonts(self) -> dict:
        fonts: dict = {}
        fonts_dir = resource_path("fonts")
        if not os.path.exists(fonts_dir):
            try:
                os.makedirs(fonts_dir)
            except OSError:
                pass
        if os.path.isdir(fonts_dir):
            for f in sorted(os.listdir(fonts_dir)):
                if f.lower().endswith((".ttf", ".otf")):
                    fonts[os.path.splitext(f)[0]] = os.path.join(fonts_dir, f)
        if not fonts:
            fonts["Default"] = "arial.ttf"
        return dict(sorted(fonts.items(), key=lambda x: x[0].lower()))

    def _get_font(self, font_name: str, size: int) -> ImageFont.FreeTypeFont:
        path = self.available_fonts.get(font_name, "arial.ttf")
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            return ImageFont.load_default()

    # ══════════════════════════════════════════════════════════════
    # COLOR HELPERS
    # ══════════════════════════════════════════════════════════════

    def _hex_to_rgb(self, color) -> tuple:
        if isinstance(color, tk.StringVar):
            color = color.get()
        if color.startswith("cmyk("):
            c, m, y, k = map(float, color[5:-1].split(","))
            return (int(255*(1-c)*(1-k)),
                    int(255*(1-m)*(1-k)),
                    int(255*(1-y)*(1-k)))
        color = color.lstrip("#")
        return tuple(int(color[i:i+2], 16) for i in (0, 2, 4))

    def _rgb_to_cmyk(self, hex_color: str) -> str:
        r, g, b = [x/255 for x in self._hex_to_rgb(hex_color)]
        k = 1 - max(r, g, b)
        if k == 1:
            return "cmyk(0.00,0.00,0.00,1.00)"
        c = (1-r-k)/(1-k); m = (1-g-k)/(1-k); y = (1-b-k)/(1-k)
        return f"cmyk({c:.2f},{m:.2f},{y:.2f},{k:.2f})"

    def _cmyk_to_hex(self, cmyk: str) -> str:
        c, m, y, k = map(float, cmyk[5:-1].split(","))
        return "#{:02x}{:02x}{:02x}".format(
            int(255*(1-c)*(1-k)),
            int(255*(1-m)*(1-k)),
            int(255*(1-y)*(1-k)))

    def choose_color(self, field: str):
        cur = self.font_settings[field]["color"].get()
        if self.color_space.get() == "RGB":
            init = cur if cur.startswith("#") else "#000000"
            chosen = colorchooser.askcolor(title=f"Color – {field}", initialcolor=init)
            if chosen[1]:
                self.font_settings[field]["color"].set(chosen[1])
        else:
            self._cmyk_picker(field, cur)
        self.update_preview(field)

    def _cmyk_picker(self, field: str, current: str):
        try:
            c_val, m_val, y_val, k_val = map(float, current[5:-1].split(","))
        except Exception:
            c_val = m_val = y_val = k_val = 0.0

        win = tk.Toplevel(self.root)
        win.title(f"CMYK – {field}")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        cvars = {ch: tk.DoubleVar(value=v)
                 for ch, v in zip("CMYK", (c_val, m_val, y_val, k_val))}
        preview = tk.Label(win, width=20, height=5)
        preview.grid(row=4, column=0, columnspan=2, pady=8)

        def refresh(*_):
            cmyk = "cmyk({:.2f},{:.2f},{:.2f},{:.2f})".format(
                cvars["C"].get(), cvars["M"].get(),
                cvars["Y"].get(), cvars["K"].get())
            self.font_settings[field]["color"].set(cmyk)
            preview.config(bg=self._cmyk_to_hex(cmyk))

        for row, (label, ch) in enumerate(zip(("Cyan","Magenta","Yellow","Black"), "CMYK")):
            tk.Label(win, text=f"{label}:").grid(row=row, column=0, padx=8, pady=4, sticky="e")
            tk.Scale(win, from_=0, to=1, resolution=0.01,
                     variable=cvars[ch], command=refresh,
                     orient="horizontal", length=200).grid(row=row, column=1, padx=8)

        refresh()
        tk.Button(win, text="OK", command=win.destroy,
                  bg=self.ACCENT, fg="white", relief="flat",
                  padx=20, pady=4).grid(row=5, column=0, columnspan=2, pady=8)
        self.root.wait_window(win)

    # ══════════════════════════════════════════════════════════════
    # TEMPLATE / EXCEL LOADING
    # ══════════════════════════════════════════════════════════════

    def load_template(self, file_path: str | None = None):
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="Select Certificate Template",
                filetypes=[("PNG Image", "*.png"), ("All Images", "*.png *.jpg *.jpeg")])
        if not file_path:
            return
        try:
            self.template_path  = file_path
            self.original_image = Image.open(file_path).convert("RGBA")
            self._refresh_canvas()
            for field in self.fields:
                self.create_placeholder(field)
            self._update_status(f"Template loaded: {os.path.basename(file_path)}")
        except Exception as exc:
            messagebox.showerror("Error", f"Cannot load template:\n{exc}")

    def _refresh_canvas(self):
        """Resize image to fit canvas and redraw background."""
        if self.original_image is None:
            return
        ow, oh = self.original_image.size
        mw, mh = 1000, 700
        ratio   = min(mw/ow, mh/oh)
        nw, nh  = int(ow*ratio), int(oh*ratio)
        self.scale_x, self.scale_y = ow/nw, oh/nh
        resized = self.original_image.resize((nw, nh), Image.LANCZOS)
        self.display_image = ImageTk.PhotoImage(resized)
        self.canvas.config(width=nw, height=nh)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, image=self.display_image, anchor="nw")
        # Redraw existing placeholders
        for field in list(self.placeholders.keys()):
            x, y = self.placeholders[field]["x"], self.placeholders[field]["y"]
            self._draw_placeholder(field, x, y)

    def load_excel(self, file_path: str | None = None):
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        try:
            wb    = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            header = [str(c.value).strip().lower()
                      for c in next(sheet.iter_rows(min_row=1, max_row=1))
                      if c.value is not None]
            if not header:
                messagebox.showerror("Error", "No headers found in Excel file.")
                return

            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i >= len(header):
                        break
                    if isinstance(val, datetime):
                        val = val.strftime("%d-%m-%Y")
                    rec[header[i]] = str(val) if val is not None else ""
                if any(rec.values()):
                    data.append(rec)
            wb.close()

            if not data:
                messagebox.showwarning("Warning", "No data rows found in the Excel file.")
                return

            self.excel_path = file_path
            self.fields     = header
            self.excel_data = data

            # Init per-field settings
            default_font = next(iter(self.available_fonts))
            self.field_vars    = {f: tk.BooleanVar(value=True) for f in header}
            self.font_settings = {
                f: {
                    "size":      tk.IntVar(value=32),
                    "color":     tk.StringVar(value="#000000"),
                    "font_name": tk.StringVar(value=default_font),
                } for f in header}

            self._build_field_panel()
            if self.original_image:
                for f in self.fields:
                    self.create_placeholder(f)

            messagebox.showinfo("Loaded", f"{len(data)} records loaded successfully.")
            self._update_status(f"Excel loaded – {len(data)} records, {len(header)} fields")
        except Exception as exc:
            messagebox.showerror("Error", f"Cannot load Excel file:\n{exc}")

    # ══════════════════════════════════════════════════════════════
    # FIELD PANEL
    # ══════════════════════════════════════════════════════════════

    def _build_field_panel(self):
        """Rebuild the left-panel field list."""
        for w in self.toggle_frame.winfo_children():
            w.destroy()

        inner_canvas = tk.Canvas(self.toggle_frame, height=300,
                                 bg=self.WHITE, highlightthickness=0)
        vsb = ttk.Scrollbar(self.toggle_frame, orient="vertical",
                            command=inner_canvas.yview)
        frame = tk.Frame(inner_canvas, bg=self.WHITE)
        frame.bind("<Configure>",
                   lambda e: inner_canvas.configure(
                       scrollregion=inner_canvas.bbox("all")))
        inner_canvas.create_window((0, 0), window=frame, anchor="nw")
        inner_canvas.configure(yscrollcommand=vsb.set)

        for field in self.fields:
            self._add_field_row(frame, field)

        vsb.pack(side="right", fill="y")
        inner_canvas.pack(side="left", fill="both", expand=True)

    def _add_field_row(self, parent, field: str):
        row = tk.Frame(parent, relief="solid", borderwidth=1,
                       padx=10, pady=8, bg=self.WHITE)
        row.pack(fill="x", pady=2, padx=2)

        tk.Label(row, text=field.title(),
                 font=("Arial", 10, "bold"),
                 bg=self.WHITE, anchor="w").pack(fill="x", pady=(0, 4))

        ctrl = tk.Frame(row, bg=self.WHITE)
        ctrl.pack(fill="x")

        # Size spinbox
        tk.Label(ctrl, text="Size:", bg=self.WHITE).pack(side="left")
        spin = tk.Spinbox(ctrl, from_=8, to=300,
                          textvariable=self.font_settings[field]["size"],
                          width=4, command=lambda f=field: self.update_preview(f))
        spin.bind("<Return>", lambda e, f=field: self.update_preview(f))
        spin.pack(side="left", padx=(0, 8))

        # Font combobox
        tk.Label(ctrl, text="Font:", bg=self.WHITE).pack(side="left")
        cb = ttk.Combobox(ctrl,
                          values=list(self.available_fonts.keys()),
                          textvariable=self.font_settings[field]["font_name"],
                          width=14, state="readonly")
        cb.bind("<<ComboboxSelected>>", lambda e, f=field: self.update_preview(f))
        cb.pack(side="left", padx=(0, 8))

        # Color button
        color_lbl = tk.Label(ctrl, width=3, height=1,
                             bg=self.font_settings[field]["color"].get(),
                             relief="solid", borderwidth=1)
        color_lbl.pack(side="left", padx=(0, 4))
        self.font_settings[field]["_color_lbl"] = color_lbl
        tk.Button(ctrl, text="Color",
                  command=lambda f=field: self._pick_color(f),
                  relief="flat", bg="#e8e8e8",
                  padx=8, pady=2, font=("Arial", 9)).pack(side="left")

    def _pick_color(self, field: str):
        self.choose_color(field)
        lbl = self.font_settings[field].get("_color_lbl")
        if lbl:
            col = self.font_settings[field]["color"].get()
            if col.startswith("cmyk("):
                col = self._cmyk_to_hex(col)
            try:
                lbl.config(bg=col)
            except Exception:
                pass

    # ══════════════════════════════════════════════════════════════
    # PLACEHOLDER MANAGEMENT
    # ══════════════════════════════════════════════════════════════

    def _render_placeholder_image(self, field: str):
        """Return an ImageTk.PhotoImage for the canvas placeholder."""
        size  = self.font_settings[field]["size"].get()
        color = self.font_settings[field]["color"].get()
        fname = self.font_settings[field]["font_name"].get()
        font  = self._get_font(fname, size)

        sample = (self.excel_data[0].get(field, field)
                  if self.excel_data else field) or field

        # Measure text
        tmp   = Image.new("RGBA", (1, 1))
        draw  = ImageDraw.Draw(tmp)
        tw    = int(draw.textlength(sample, font=font))
        try:
            asc, desc = font.getmetrics()
            th = asc + desc
        except Exception:
            th = size
        tw, th = max(tw, 1), max(th, 1)

        img  = Image.new("RGBA", (tw, th), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        try:
            bbox = font.getbbox(sample)
            yo   = (th - (bbox[3] - bbox[1])) // 2
        except Exception:
            yo = 0
        draw.text((0, yo), sample, font=font, fill=self._hex_to_rgb(color))

        sw = max(int(tw / self.scale_x), 1)
        sh = max(int(th / self.scale_y), 1)
        return ImageTk.PhotoImage(img.resize((sw, sh), Image.LANCZOS))

    def _draw_placeholder(self, field: str, x: float, y: float):
        """Draw / redraw a canvas item for field at canvas coords (x, y)."""
        # Remove old item if exists
        if field in self.placeholders:
            self.canvas.delete(self.placeholders[field]["item"])

        photo = self._render_placeholder_image(field)
        self._placeholder_images[field] = photo   # keep ref
        item  = self.canvas.create_image(x, y, image=photo, anchor="center")
        self.canvas.tag_bind(item, "<Button-1>",  lambda e, i=item: self._drag_start(e, i))
        self.canvas.tag_bind(item, "<B1-Motion>", lambda e, i=item: self._drag_move(e, i))
        self.placeholders[field] = {"item": item, "x": x, "y": y}

    def create_placeholder(self, field: str, x=None, y=None, **_):
        """Public helper – creates or moves a placeholder."""
        if field not in self.fields:
            return
        if x is None or y is None:
            # Default: horizontally centred, stacked vertically
            cw = self.canvas.winfo_width() or 800
            idx = self.fields.index(field)
            x   = cw // 2
            y   = 50 + idx * 60
        self._draw_placeholder(field, x, y)

    def update_preview(self, field: str):
        if field in self.placeholders:
            pos = self.placeholders[field]
            self._draw_placeholder(field, pos["x"], pos["y"])
            self._update_status(f"Updated: {field}")

    # ── drag & drop ────────────────────────────────────────────
    def _drag_start(self, event, item):
        self._drag_data = {"item": item, "x": event.x, "y": event.y}

    def _drag_move(self, event, item):
        dx = event.x - self._drag_data["x"]
        dy = event.y - self._drag_data["y"]
        self.canvas.move(item, dx, dy)
        self._drag_data["x"] = event.x
        self._drag_data["y"] = event.y
        # Update stored position
        for field, data in self.placeholders.items():
            if data["item"] == item:
                data["x"] += dx
                data["y"] += dy
                break

    # ══════════════════════════════════════════════════════════════
    # COORDINATE SCALING
    # ══════════════════════════════════════════════════════════════

    def _get_scaled_positions(self) -> dict:
        coords = {}
        for field, data in self.placeholders.items():
            coords[field] = (data["x"] * self.scale_x,
                             data["y"] * self.scale_y)
        return coords

    # ══════════════════════════════════════════════════════════════
    # TEXT DRAWING (shared by preview + generate)
    # ══════════════════════════════════════════════════════════════

    def _draw_text_on_image(self, img: Image.Image,
                            student: dict, positions: dict) -> Image.Image:
        draw = ImageDraw.Draw(img)
        for field in self.fields:
            if not self.field_vars.get(field, tk.BooleanVar(value=False)).get():
                continue
            if field not in positions:
                continue
            try:
                x, y    = positions[field]
                size    = self.font_settings[field]["size"].get()
                color   = self.font_settings[field]["color"].get()
                fname   = self.font_settings[field]["font_name"].get()
                font    = self._get_font(fname, size)
                text    = student.get(field, "")
                tw      = draw.textlength(text, font=font)
                try:
                    bbox    = font.getbbox(text)
                    th      = bbox[3] - bbox[1]
                    yo      = (size - th) // 2
                except Exception:
                    yo = 0
                draw.text((x - tw/2, y - size/2 + yo),
                           text, font=font, fill=self._hex_to_rgb(color))
            except Exception as exc:
                print(f"[draw_text] {field}: {exc}")
        return img

    # ══════════════════════════════════════════════════════════════
    # PREVIEW
    # ══════════════════════════════════════════════════════════════

    def preview_certificate(self):
        if not self.original_image:
            messagebox.showwarning("CertWizard", "Please load a template first.")
            return
        if not self.excel_data:
            messagebox.showwarning("CertWizard", "Please load student data first.")
            return

        img = self._draw_text_on_image(
            self.original_image.copy(), self.excel_data[0],
            self._get_scaled_positions())

        win = tk.Toplevel(self.root)
        win.title("CertWizard – Preview")
        win.transient(self.root)
        win.grab_set()
        win.resizable(True, True)
        win.minsize(400, 300)

        pw, ph = max(self.root.winfo_width()-100, 600), \
                 max(self.root.winfo_height()-100, 400)
        win.geometry(f"{pw}x{ph}")

        iw, ih   = img.size
        ratio    = min(pw/iw, ph/ih)
        prev_img = img.resize((int(iw*ratio), int(ih*ratio)), Image.LANCZOS)
        photo    = ImageTk.PhotoImage(prev_img)

        lbl = tk.Label(win, image=photo, bg=self.WHITE)
        lbl.image = photo
        lbl.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Button(win, text="Close", command=win.destroy,
                  bg=self.ACCENT, fg="white", relief="flat",
                  padx=20, pady=5, font=("Arial", 10, "bold"),
                  activebackground="#2980b9").pack(pady=8)
        win.bind("<Escape>", lambda e: win.destroy())
        win.lift(); win.focus_set()

    # ══════════════════════════════════════════════════════════════
    # GENERATION
    # ══════════════════════════════════════════════════════════════

    def generate_certificates(self):
        if not self.excel_data:
            messagebox.showwarning("CertWizard", "Please load student data first.")
            return
        if not self.original_image:
            messagebox.showwarning("CertWizard", "Please load a template first.")
            return
        if not self._gen_lock.acquire(blocking=False):
            messagebox.showwarning("CertWizard", "Generation already in progress.")
            return

        use_cmyk = messagebox.askyesno(
            "Color Space",
            "Generate certificates in CMYK color space?\n\nYes = CMYK    No = RGB")
        self.color_space.set("CMYK" if use_cmyk else "RGB")

        out_dir = filedialog.askdirectory(title="Select Output Folder")
        if not out_dir:
            self._gen_lock.release()
            return

        sub = "CMYK" if use_cmyk else "RGB"
        os.makedirs(os.path.join(out_dir, sub), exist_ok=True)

        positions = self._get_scaled_positions()
        total     = len(self.excel_data)

        def px_to_mm(px): return px * 0.264583
        iw, ih   = self.original_image.size
        pdf_w    = px_to_mm(iw)
        pdf_h    = px_to_mm(ih)

        import tempfile

        def _run():
            count = 0
            self._log("Starting generation...", clear=True)
            self._log(f"Total: {total}  |  Color: {sub}  |  Out: {out_dir}")
            self._log("-" * 40)

            for idx, student in enumerate(self.excel_data):
                try:
                    img  = self._draw_text_on_image(
                        self.original_image.copy().convert("RGB"), student, positions)
                    pdf  = FPDF(unit="mm", format=(pdf_w, pdf_h))
                    pdf.add_page()

                    with tempfile.NamedTemporaryFile(
                            suffix=".png", delete=False) as tmp:
                        tmp_path = tmp.name
                    try:
                        img.save(tmp_path, format="PNG", optimize=True)
                        pdf.image(tmp_path, x=0, y=0, w=pdf_w, h=pdf_h)
                    finally:
                        try:
                            os.remove(tmp_path)
                        except OSError:
                            pass

                    # Safe filename from first two fields
                    parts = [re.sub(r'[^\w\-_. ]', '',
                                    str(student.get(self.fields[i], ""))).strip()
                             for i in range(min(2, len(self.fields)))]
                    safe  = "_".join(p for p in parts if p) or f"cert_{idx+1}"
                    out   = os.path.join(out_dir, sub, f"{safe}_certificate.pdf")
                    pdf.output(out)
                    count += 1
                    self._log(f"[{idx+1}/{total}] {safe}_certificate.pdf")
                except Exception as exc:
                    self._log(f"[ERROR] cert {idx+1}: {exc}")

                pct = (idx+1) / total * 100
                self.root.after(0, lambda v=pct: self.progress.configure(value=v))

            self._log("-" * 40)
            self._log(f"Done – {count}/{total} certificates generated.")
            self.root.after(0, lambda: messagebox.showinfo(
                "CertWizard", f"{count} certificate(s) generated!"))
            self._gen_lock.release()

        threading.Thread(target=_run, daemon=True).start()

    # ══════════════════════════════════════════════════════════════
    # PROJECT SAVE / LOAD
    # ══════════════════════════════════════════════════════════════

    def save_project(self):
        if not self.original_image:
            messagebox.showwarning("Warning", "No template loaded.")
            return
        try:
            field_settings = {}
            for f in self.fields:
                field_settings[f] = {
                    "size":      self.font_settings[f]["size"].get(),
                    "color":     self.font_settings[f]["color"].get(),
                    "visible":   self.field_vars[f].get(),
                    "font_name": self.font_settings[f]["font_name"].get(),
                }
            project = {
                "version":       "2.0",
                "last_modified": datetime.now().isoformat(),
                "template_path": self.template_path,
                "excel_path":    self.excel_path,
                "color_space":   self.color_space.get(),
                "positions":     self._get_scaled_positions(),
                "field_settings": field_settings,
            }
            path = filedialog.asksaveasfilename(
                defaultextension=".certwiz",
                filetypes=[("CertWizard Project", "*.certwiz")])
            if path:
                with open(path, "w", encoding="utf-8") as fh:
                    json.dump(project, fh, indent=2)
                messagebox.showinfo("Saved", "Project saved successfully.")
        except Exception as exc:
            messagebox.showerror("Error", f"Save failed:\n{exc}")

    def load_project(self):
        path = filedialog.askopenfilename(
            filetypes=[("CertWizard Project", "*.certwiz")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)

            self.canvas.delete("all")
            self.placeholders.clear()
            self._placeholder_images.clear()

            tpl = data.get("template_path", "")
            if tpl and os.path.exists(tpl):
                self.load_template(tpl)
            elif tpl:
                messagebox.showwarning("Warning",
                    f"Template not found at saved path:\n{tpl}")

            xl = data.get("excel_path", "")
            if xl and os.path.exists(xl):
                self.load_excel(xl)
            elif xl:
                messagebox.showwarning("Warning",
                    f"Excel file not found at saved path:\n{xl}")

            self.color_space.set(data.get("color_space", "RGB"))

            fs = data.get("field_settings", {})
            for f in self.fields:
                if f in fs:
                    self.font_settings[f]["size"].set(fs[f].get("size", 32))
                    self.font_settings[f]["color"].set(fs[f].get("color", "#000000"))
                    self.font_settings[f]["font_name"].set(
                        fs[f].get("font_name", next(iter(self.available_fonts))))
                    self.field_vars[f].set(fs[f].get("visible", True))

            for field, (sx, sy) in data.get("positions", {}).items():
                if field in self.fields:
                    cx = sx / self.scale_x
                    cy = sy / self.scale_y
                    self._draw_placeholder(field, cx, cy)

            messagebox.showinfo("Loaded", "Project loaded successfully.")
        except Exception as exc:
            messagebox.showerror("Error", f"Load failed:\n{exc}")

    # ══════════════════════════════════════════════════════════════
    # STATUS / LOG HELPERS
    # ══════════════════════════════════════════════════════════════

    def _update_status(self, msg: str):
        self.status_bar.config(text=f"CertWizard: {msg}")
        self.root.update_idletasks()

    def _log(self, msg: str, clear: bool = False):
        def _inner():
            self.info_text.configure(state="normal")
            if clear:
                self.info_text.delete("1.0", tk.END)
            self.info_text.insert(tk.END, msg + "\n")
            self.info_text.see(tk.END)
            self.info_text.configure(state="disabled")
        self.root.after(0, _inner)

    # ── legacy aliases so old .certwiz / tests don't break ─────
    update_status = _update_status
    update_info   = lambda self, m, clear=False: self._log(m, clear)
    get_placeholder_positions = _get_scaled_positions


if __name__ == "__main__":
    root = tk.Tk()
    app  = CertificateApp(root)
    root.mainloop()
